"""Excel workbook building and parsing for the product bulk export/import.

One template serves both directions: what the export writes is exactly what the
import reads back. Column order is irrelevant on import — headers are matched by
name — but a missing required header rejects the file outright.
"""

import io
import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_VERSION = "1"

DATA_SHEET = "Products"
META_SHEET = "_meta"
REFERENCE_SHEET = "Reference"

COLUMNS = [
    "id",
    "code",
    "sku",
    "name",
    "description",
    "type",
    "family",
    "category",
    "size",
    "material_type",
    "specification",
    "manufacturer",
    "model",
    "unit_price",
    "cost",
    "tax_rate",
    "unit",
    "is_active",
]

# Without these the file cannot be interpreted at all.
REQUIRED_COLUMNS = {"name", "type"}

COLUMN_WIDTHS = {
    "id": 8,
    "code": 16,
    "sku": 16,
    "name": 40,
    "description": 40,
    "type": 12,
    "family": 20,
    "category": 20,
    "unit_price": 12,
    "cost": 12,
    "tax_rate": 10,
}

HEADER_FILL = PatternFill(start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid")

PRODUCT_TYPES = ["Product", "Service"]
BOOLEAN_VALUES = ["TRUE", "FALSE"]
# Dropdowns cover this many rows past the data so appended rows keep them.
DROPDOWN_HEADROOM = 500


class InvalidWorkbookError(Exception):
    """The uploaded file is not a workbook this importer can read."""


@dataclass
class ParsedRow:
    row_number: int
    values: dict[str, Any]


@dataclass
class ParsedFile:
    rows: list[ParsedRow] = field(default_factory=list)
    exported_at: datetime | None = None
    export_scope: str | None = None


def build_export_workbook(
    products: list[Any],
    *,
    exported_at: datetime,
    export_scope: dict[str, Any],
    families: list[Any],
    categories: list[tuple[str, str]],
) -> Workbook:
    """Build the round-trippable export workbook.

    `products` are ProductRead-shaped objects; `categories` are (family, category)
    pairs so the reference sheet shows which family each category belongs to.
    """
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = DATA_SHEET
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    sheet.freeze_panes = "A2"

    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 18)

    for product in products:
        sheet.append(
            [
                product.id,
                _cell(product.code),
                _cell(product.sku),
                _cell(product.name),
                _cell(product.description),
                _cell(product.type),
                _cell(product.family_name),
                _cell(product.category_name),
                _cell(product.size),
                _cell(product.material_type),
                _cell(product.specification),
                _cell(product.manufacturer),
                _cell(product.model),
                float(product.unit_price),
                float(product.cost),
                float(product.tax_rate),
                product.unit,
                # Written as text, not a bool: Excel localises real booleans
                # ("VERDADERO" in Spanish) and the file stops round-tripping.
                "TRUE" if product.is_active else "FALSE",
            ]
        )

    # The id is what lets a rename keep pointing at the same product, but it is
    # noise for whoever edits the sheet, so it stays out of sight.
    sheet.column_dimensions[get_column_letter(COLUMNS.index("id") + 1)].hidden = True

    _add_dropdown(sheet, "type", PRODUCT_TYPES, len(products))
    _add_dropdown(sheet, "is_active", BOOLEAN_VALUES, len(products))

    meta = workbook.create_sheet(META_SHEET)
    meta.append(["template_version", TEMPLATE_VERSION])
    meta.append(["exported_at", exported_at.isoformat()])
    meta.append(["export_scope", json.dumps(export_scope)])
    # Hidden, not protected: it carries machine data, not something to edit.
    meta.sheet_state = "hidden"

    reference = workbook.create_sheet(REFERENCE_SHEET)
    reference.append(["Valid values"])
    reference.append(["type", "Product", "Service"])
    reference.append(["is_active", "TRUE", "FALSE"])
    reference.append([])
    reference.append(["Families"])
    for family in families:
        reference.append([family.name])
    reference.append([])
    reference.append(["Family", "Category"])
    for family_name, category_name in categories:
        reference.append([family_name, category_name])
    reference.column_dimensions["A"].width = 28
    reference.column_dimensions["B"].width = 28

    return workbook


def _cell(value: Any) -> Any:
    r"""Write text with `\n` line endings only.

    A cell lives inside XML, and XML parsers must normalise carriage returns. A
    stored "\r\n" therefore comes back as "\n\n" — the text gains a blank line on
    every round trip and looks edited when nobody touched it.
    """
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n")
    return value


def _add_dropdown(sheet, column: str, options: list[str], row_count: int) -> None:
    """Constrain a column to a fixed list, so nobody has to guess the spelling."""
    letter = get_column_letter(COLUMNS.index(column) + 1)
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error=f"Choose one of: {', '.join(options)}",
    )
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{row_count + 1 + DROPDOWN_HEADROOM}")


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def count_data_rows(path: str) -> int:
    """Count data rows without materialising them, to enforce the row limit early."""
    workbook = _load(path)
    try:
        sheet = _data_sheet(workbook)
        total = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(value is not None and str(value).strip() for value in row):
                total += 1
        return total
    finally:
        workbook.close()


def parse_import_workbook(path: str) -> ParsedFile:
    workbook = _load(path)
    try:
        sheet = _data_sheet(workbook)
        header = _read_header(sheet)

        parsed = ParsedFile(**_read_meta(workbook))
        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            values = {column: row[index] if index < len(row) else None for column, index in header.items()}
            parsed.rows.append(ParsedRow(row_number=excel_row, values=values))
        return parsed
    finally:
        workbook.close()


def build_error_report(rows: list[Any]) -> bytes:
    """Workbook listing the rows that could not be applied, with the reason."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Errors"
    sheet.append(["Excel row", "Code", "Name", "Reason"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in rows:
        sheet.append([row.row_number, row.code, row.name, row.message])

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 40
    sheet.column_dimensions["D"].width = 60
    return workbook_to_bytes(workbook)


def _load(path: str) -> Workbook:
    size = Path(path).stat().st_size
    if size == 0:
        raise InvalidWorkbookError("The uploaded file is empty. Export the catalog again.")

    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise InvalidWorkbookError(_unreadable_reason(path, size, exc)) from exc


def _unreadable_reason(path: str, size: int, exc: Exception) -> str:
    """Name what the file actually is, instead of leaking zip jargon.

    An .xlsx is a zip whose central directory sits at the end, so a download that
    was cut short still starts with `PK` but cannot be opened at all. That is the
    most common failure here and it is not recoverable by editing the file.
    """
    raw = Path(path).read_bytes()
    head = raw[:4]

    if head.startswith(b"PK"):
        return (
            f"This .xlsx is incomplete ({size} bytes) — the download was cut short, so the file "
            "on your computer is only part of the spreadsheet. Export the catalog again and "
            "upload the new file."
        )
    if head.startswith(b"<"):
        # Excel Online saves its own error page under the .xlsx name when it
        # fails to open a workbook, which is indistinguishable from the file
        # itself until you look inside.
        if b"ewaResult" in raw[:4096] or b"InternalErrorEwr" in raw[:4096]:
            return (
                "This is an Excel Online error page, not a spreadsheet. Excel Online failed to "
                "open the workbook and saved its error under the .xlsx name. Export the catalog "
                "again and open it with the Excel desktop app, or upload it without opening it."
            )
        return f"This file is a web page, not a spreadsheet ({size} bytes). Export the catalog again."
    if head[:1] in {b"{", b"["}:
        return f"This file is a server response, not a spreadsheet ({size} bytes). Export the catalog again."
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return (
            "This is a legacy .xls file. Open it in Excel and save it as .xlsx "
            "(Excel Workbook) before uploading."
        )
    return f"The file could not be read as an Excel workbook ({size} bytes): {exc}"


def _data_sheet(workbook: Workbook):
    if DATA_SHEET in workbook.sheetnames:
        return workbook[DATA_SHEET]
    # A hand-built file may not have named the sheet; fall back to the first one.
    return workbook.worksheets[0]


def _read_header(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise InvalidWorkbookError("The sheet has no header row.")

    header: dict[str, int] = {}
    for index, value in enumerate(header_row):
        if value is None:
            continue
        name = str(value).strip().lower().replace(" ", "_")
        if name in COLUMNS and name not in header:
            header[name] = index

    missing = REQUIRED_COLUMNS - header.keys()
    if missing:
        raise InvalidWorkbookError(f"Missing required column(s): {', '.join(sorted(missing))}.")
    return header


def _read_meta(workbook: Workbook) -> dict[str, Any]:
    if META_SHEET not in workbook.sheetnames:
        return {}

    meta: dict[str, str] = {}
    for row in workbook[META_SHEET].iter_rows(min_row=1, max_col=2, values_only=True):
        if row and row[0]:
            meta[str(row[0]).strip()] = str(row[1]) if row[1] is not None else ""

    exported_at = None
    raw_exported_at = meta.get("exported_at")
    if raw_exported_at:
        try:
            exported_at = datetime.fromisoformat(raw_exported_at)
        except ValueError:
            exported_at = None

    return {"exported_at": exported_at, "export_scope": meta.get("export_scope")}
