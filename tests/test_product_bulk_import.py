import io
from datetime import timedelta

import pytest
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

import api.product_bulk_imports as bulk_api
from core.datetime_utils import utcnow
from models.product_bulk_imports import ProductBulkImportRows, ProductBulkImports
from models.products import ProductCategories, ProductFamilies, Products
from services.products.bulk_import import run_analysis, run_apply
from services.products.excel import COLUMNS

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(autouse=True)
def _grant_bulk_permissions(permission_override) -> None:
    permission_override("products", {"can_view", "can_create", "can_edit", "can_export"})


@pytest.fixture(autouse=True)
def _isolate_upload_dir(tmp_path, monkeypatch) -> None:
    upload_dir = tmp_path / "bulk-imports"
    upload_dir.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(bulk_api, "UPLOAD_DIR", upload_dir)


@pytest.fixture
def session_factory(db_session: Session):
    """Background phases open their own session; in tests they reuse the test one."""

    class _NonClosing:
        def __init__(self, session: Session) -> None:
            self.session = session

        def __enter__(self) -> Session:
            return self.session

        def __exit__(self, *_args) -> None:
            return None

    return lambda: _NonClosing(db_session)


@pytest.fixture
def employee(db_session: Session):
    from models.employees import Employees

    emp = db_session.exec(select(Employees).where(Employees.email == "test@example.com")).first()
    if not emp:
        emp = Employees(employee_id=1, email="test@example.com", display_name="Test User")
        db_session.add(emp)
        db_session.commit()
    return emp


@pytest.fixture
def taxonomy(db_session: Session):
    family = ProductFamilies(name="Alarms")
    db_session.add(family)
    db_session.commit()
    category = ProductCategories(family_id=family.id, name="Panels")
    db_session.add(category)
    db_session.commit()
    return family, category


def build_file(rows: list[dict], *, exported_at=None, columns: list[str] | None = None) -> bytes:
    """Build an import workbook from dicts keyed by template column name."""
    columns = columns or COLUMNS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])

    if exported_at is not None:
        meta = workbook.create_sheet("_meta")
        meta.append(["template_version", "1"])
        meta.append(["exported_at", exported_at.isoformat()])
        meta.append(["export_scope", "{}"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_product(db: Session, **kwargs) -> Products:
    defaults = {"name": "Panel", "type": "Product", "unit": "pieza"}
    defaults.update(kwargs)
    product = Products(**defaults)
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


def create_job(db: Session, content: bytes, tmp_path, **kwargs) -> ProductBulkImports:
    path = tmp_path / f"job_{len(list(tmp_path.iterdir()))}.xlsx"
    path.write_bytes(content)
    kwargs.setdefault("status", "analyzing")
    job = ProductBulkImports(
        file_name="import.xlsx",
        stored_path=str(path),
        created_by=1,
        **kwargs,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def analyze(db: Session, session_factory, content: bytes, tmp_path) -> ProductBulkImports:
    job = create_job(db, content, tmp_path)
    run_analysis(job.id, session_factory)
    db.refresh(job)
    return job


def apply_job(db: Session, session_factory, job: ProductBulkImports, **confirmation) -> ProductBulkImports:
    job.status = "applying"
    for key, value in confirmation.items():
        setattr(job, key, value)
    db.commit()
    run_apply(job.id, session_factory)
    db.refresh(job)
    return job


def rows_of(db: Session, job: ProductBulkImports) -> list[ProductBulkImportRows]:
    return db.exec(
        select(ProductBulkImportRows)
        .where(ProductBulkImportRows.import_id == job.id)
        .order_by(ProductBulkImportRows.row_number)
    ).all()


# ----------------------------
# Export
# ----------------------------
def test_export_returns_xlsx_with_all_filtered_products(client, auth_headers, db_session, taxonomy):
    family, category = taxonomy
    make_product(db_session, name="Panel A", code="PA-1", family_id=family.id, category_id=category.id)
    make_product(db_session, name="Panel B", code="PB-1", family_id=family.id, category_id=category.id)

    response = client.get("/products/export", headers=auth_headers)

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE

    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook["Products"].max_row == 3
    assert [cell.value for cell in workbook["Products"][1]] == COLUMNS
    assert "_meta" in workbook.sheetnames


def test_export_respects_filters(client, auth_headers, db_session, taxonomy):
    family, _ = taxonomy
    make_product(db_session, name="Included", family_id=family.id)
    make_product(db_session, name="Excluded")

    response = client.get(f"/products/export?family_id={family.id}", headers=auth_headers)

    workbook = load_workbook(io.BytesIO(response.content))
    names = [row[3] for row in workbook["Products"].iter_rows(min_row=2, values_only=True)]
    assert names == ["Included"]


def test_export_writes_is_active_as_text_not_a_localised_boolean(
    client, auth_headers, db_session
):
    """A real boolean renders as VERDADERO/FALSO in a Spanish Excel and stops round-tripping."""
    make_product(db_session, name="Active One", is_active=True)
    make_product(db_session, name="Inactive One", is_active=False)

    response = client.get("/products/export", headers=auth_headers)

    sheet = load_workbook(io.BytesIO(response.content))["Products"]
    column = COLUMNS.index("is_active")
    values = [row[column] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert sorted(values) == ["FALSE", "TRUE"]


def test_export_constrains_type_and_status_with_dropdowns(client, auth_headers, db_session):
    make_product(db_session, name="Panel")

    response = client.get("/products/export", headers=auth_headers)

    sheet = load_workbook(io.BytesIO(response.content))["Products"]
    formulas = {validation.formula1 for validation in sheet.data_validations.dataValidation}
    assert '"Product,Service"' in formulas
    assert '"TRUE,FALSE"' in formulas


def test_export_hides_the_id_column(client, auth_headers, db_session):
    make_product(db_session, name="Panel")

    response = client.get("/products/export", headers=auth_headers)

    sheet = load_workbook(io.BytesIO(response.content))["Products"]
    assert sheet.column_dimensions["A"].hidden is True


def test_localised_boolean_is_still_accepted_on_import(db_session, session_factory, tmp_path):
    product = make_product(db_session, name="Panel", code="L-1")
    content = build_file(
        [{"id": product.id, "code": "L-1", "name": "Panel", "type": "Product", "is_active": "VERDADERO"}]
    )

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 0


def test_export_requires_can_export(client, auth_headers, permission_override):
    permission_override("products", {"can_view"})

    response = client.get("/products/export", headers=auth_headers)

    assert response.status_code == 403


# ----------------------------
# Analysis
# ----------------------------
def test_round_trip_without_edits_reports_everything_unchanged(
    db_session, session_factory, tmp_path, taxonomy
):
    family, category = taxonomy
    product = make_product(
        db_session,
        name="Panel A",
        code="PA-1",
        family_id=family.id,
        category_id=category.id,
    )
    content = build_file(
        [
            {
                "id": product.id,
                "code": "PA-1",
                "name": "Panel A",
                "type": "Product",
                "family": "Alarms",
                "category": "Panels",
                "unit": "pieza",
                "unit_price": 0,
                "cost": 0,
                "tax_rate": 0,
                "is_active": True,
            }
        ]
    )

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.status == "awaiting_confirmation"
    assert job.unchanged_count == 1
    assert job.create_count == 0
    assert job.update_count == 0
    assert rows_of(db_session, job) == []


def test_round_trip_does_not_invent_updates_from_messy_stored_data(
    client, auth_headers, db_session, session_factory, tmp_path, taxonomy
):
    """Normalising the file but not the stored value turns tidy-ups into fake edits."""
    from decimal import Decimal

    family, _ = taxonomy
    # Exactly the shapes a real catalog holds: padded code, lowercase type,
    # blank unit, trailing-zero decimals.
    make_product(db_session, name="Padded code", code="B300-6 ", unit="pieza")
    make_product(db_session, name="Lowercase type", code="L-1", type="product", unit="pieza")
    make_product(db_session, name="Blank unit", code="E-1", unit="")
    make_product(
        db_session,
        name="Trailing zeros",
        code="Z-1",
        unit="pieza",
        unit_price=Decimal("17.50"),
        tax_rate=Decimal("11.50"),
        family_id=family.id,
    )

    exported = client.get("/products/export", headers=auth_headers)
    assert exported.status_code == 200, exported.text

    job = analyze(db_session, session_factory, exported.content, tmp_path)

    assert job.update_count == 0, [row.message for row in rows_of(db_session, job)]
    assert job.unchanged_count == 4
    assert job.create_count == 0
    assert job.error_count == 0


def test_windows_line_endings_survive_the_round_trip(
    client, auth_headers, db_session, session_factory, tmp_path
):
    r"""XML must normalise "\r", so a stored "\r\n" would come back as "\n\n"."""
    stored = "Package includes:\r\n• A website\r\n• A domain\r\n• An SSL certificate"
    make_product(db_session, name="Webpages", code="W-1", description=stored)

    exported = client.get("/products/export", headers=auth_headers)
    assert exported.status_code == 200, exported.text

    sheet = load_workbook(io.BytesIO(exported.content))["Products"]
    cell = next(sheet.iter_rows(min_row=2, values_only=True))[COLUMNS.index("description")]
    assert "\n\n" not in cell
    assert cell.count("\n") == 3

    job = analyze(db_session, session_factory, exported.content, tmp_path)

    assert job.update_count == 0, [row.message for row in rows_of(db_session, job)]
    assert job.unchanged_count == 1


def test_update_rows_spell_out_what_changes(db_session, session_factory, tmp_path):
    product = make_product(db_session, name="Old Name", code="W-1", unit="pieza")

    content = build_file(
        [{"id": product.id, "code": "W-1", "name": "New Name", "type": "Product", "unit_price": 42}]
    )
    job = analyze(db_session, session_factory, content, tmp_path)

    message = rows_of(db_session, job)[0].message
    assert "name: Old Name -> New Name" in message
    assert "unit_price" in message


def test_row_missing_from_file_leaves_the_product_untouched(
    db_session, session_factory, tmp_path
):
    kept = make_product(db_session, name="Kept", code="KEEP-1")
    make_product(db_session, name="Deleted from file", code="GONE-1")

    content = build_file([{"id": kept.id, "code": "KEEP-1", "name": "Kept renamed", "type": "Product"}])
    job = analyze(db_session, session_factory, content, tmp_path)
    apply_job(db_session, session_factory, job)

    survivor = db_session.exec(select(Products).where(Products.code == "GONE-1")).first()
    assert survivor is not None
    assert survivor.is_active is True
    assert job.updated_count == 1


def test_new_row_without_id_is_created(db_session, session_factory, tmp_path):
    content = build_file([{"name": "Brand New", "type": "Service", "unit_price": 25.5}])

    job = analyze(db_session, session_factory, content, tmp_path)
    assert job.create_count == 1

    apply_job(db_session, session_factory, job)

    created = db_session.exec(select(Products).where(Products.name == "Brand New")).first()
    assert created is not None
    assert created.type == "Service"
    assert float(created.unit_price) == pytest.approx(25.5)


def test_row_without_id_matches_existing_code(db_session, session_factory, tmp_path):
    product = make_product(db_session, name="Old Name", code="MATCH-1")

    content = build_file([{"code": "MATCH-1", "name": "New Name", "type": "Product"}])
    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.update_count == 1
    assert rows_of(db_session, job)[0].product_id == product.id


def test_duplicate_code_inside_the_file_fails_every_involved_row(
    db_session, session_factory, tmp_path
):
    content = build_file(
        [
            {"name": "First", "type": "Product", "code": "DUP-1"},
            {"name": "Second", "type": "Product", "code": "DUP-1"},
        ]
    )

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 2
    assert all("appears more than once" in row.message for row in rows_of(db_session, job))


def test_ambiguous_existing_code_is_rejected(db_session, session_factory, tmp_path):
    make_product(db_session, name="One", code="AMB-1")
    make_product(db_session, name="Two", code="AMB-1")

    content = build_file([{"name": "Which one?", "type": "Product", "code": "AMB-1"}])
    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 1
    assert "matches 2 products" in rows_of(db_session, job)[0].message


def test_stale_id_falls_back_to_the_code(db_session, session_factory, tmp_path):
    """The id column is hidden, so a stale one cannot be the user's problem."""
    product = make_product(db_session, name="Old Name", code="S-9")

    content = build_file([{"id": 9999, "code": "S-9", "name": "New Name", "type": "Product"}])
    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 0
    assert job.update_count == 1
    row = rows_of(db_session, job)[0]
    assert row.product_id == product.id
    assert "matched by code instead" in row.message


def test_stale_id_without_a_code_match_creates_the_product(
    db_session, session_factory, tmp_path
):
    content = build_file([{"id": 9999, "code": "BRAND-NEW", "name": "Ghost", "type": "Product"}])

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 0
    assert job.create_count == 1

    apply_job(db_session, session_factory, job)
    assert db_session.exec(select(Products).where(Products.code == "BRAND-NEW")).first()


def test_unreadable_id_is_ignored_rather_than_blocking_the_row(
    db_session, session_factory, tmp_path
):
    content = build_file([{"id": "abc", "code": "U-1", "name": "Fine", "type": "Product"}])

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.error_count == 0
    assert job.create_count == 1
    assert "Ignored unreadable id" in rows_of(db_session, job)[0].message


@pytest.mark.parametrize(
    ("values", "expected"),
    [
        ({"name": "", "type": "Product"}, "name is required"),
        ({"name": "X", "type": "Widget"}, "type must be"),
        ({"name": "X", "type": "Product", "tax_rate": 150}, "tax_rate must be between"),
        ({"name": "X", "type": "Product", "unit_price": -1}, "cannot be negative"),
        ({"name": "X", "type": "Product", "cost": "abc"}, "not a valid number"),
        ({"name": "X", "type": "Product", "category": "Panels"}, "category requires a family"),
    ],
)
def test_invalid_rows_are_reported(db_session, session_factory, tmp_path, values, expected):
    job = analyze(db_session, session_factory, build_file([values]), tmp_path)

    assert job.error_count == 1
    assert expected in rows_of(db_session, job)[0].message


def test_one_bad_row_does_not_stop_the_rest(db_session, session_factory, tmp_path):
    content = build_file(
        [
            {"name": "Good One", "type": "Product"},
            {"name": "Bad One", "type": "Product", "tax_rate": 150},
            {"name": "Good Two", "type": "Service"},
        ]
    )

    job = analyze(db_session, session_factory, content, tmp_path)
    apply_job(db_session, session_factory, job)

    assert job.created_count == 2
    assert job.failed_count == 1
    assert db_session.exec(select(Products).where(Products.name == "Good Two")).first() is not None


# ----------------------------
# Taxonomy
# ----------------------------
def test_unknown_family_is_surfaced_in_the_preview(db_session, session_factory, tmp_path):
    content = build_file([{"name": "P", "type": "Product", "family": "Suppression"}])

    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.status == "awaiting_confirmation"
    assert '"Suppression"' in job.unknown_taxonomy
    assert job.create_count == 1


def test_unknown_family_without_consent_fails_the_row(db_session, session_factory, tmp_path):
    content = build_file([{"name": "P", "type": "Product", "family": "Suppression"}])
    job = analyze(db_session, session_factory, content, tmp_path)

    apply_job(db_session, session_factory, job, create_missing_taxonomy=False)

    assert job.failed_count == 1
    assert job.created_count == 0
    assert db_session.exec(select(ProductFamilies).where(ProductFamilies.name == "Suppression")).first() is None


def test_unknown_family_with_consent_is_created(db_session, session_factory, tmp_path):
    content = build_file(
        [{"name": "P", "type": "Product", "family": "Suppression", "category": "Sprinklers"}]
    )
    job = analyze(db_session, session_factory, content, tmp_path)

    apply_job(db_session, session_factory, job, create_missing_taxonomy=True)

    family = db_session.exec(select(ProductFamilies).where(ProductFamilies.name == "Suppression")).first()
    assert family is not None
    category = db_session.exec(
        select(ProductCategories).where(ProductCategories.name == "Sprinklers")
    ).first()
    assert category is not None
    assert category.family_id == family.id

    product = db_session.exec(select(Products).where(Products.name == "P")).first()
    assert product.family_id == family.id
    assert product.category_id == category.id
    assert job.created_count == 1


# ----------------------------
# Conflicts
# ----------------------------
def test_product_edited_after_export_is_flagged_as_conflict(db_session, session_factory, tmp_path):
    exported_at = utcnow()
    product = make_product(db_session, name="Original", code="C-1")
    product.updated_at = exported_at + timedelta(minutes=5)
    db_session.commit()

    content = build_file(
        [{"id": product.id, "code": "C-1", "name": "From file", "type": "Product"}],
        exported_at=exported_at,
    )
    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.conflict_count == 1


def test_an_unchanged_row_is_never_reported_as_a_conflict(
    db_session, session_factory, tmp_path
):
    """Nothing gets overwritten by a row that writes nothing, however stale the file."""
    exported_at = utcnow()
    product = make_product(db_session, name="Same", code="C-2", unit="pieza")
    product.updated_at = exported_at + timedelta(hours=1)
    db_session.commit()

    content = build_file(
        [{"id": product.id, "code": "C-2", "name": "Same", "type": "Product", "unit": "pieza"}],
        exported_at=exported_at,
    )
    job = analyze(db_session, session_factory, content, tmp_path)

    assert job.unchanged_count == 1
    assert job.update_count == 0
    assert job.conflict_count == 0


def test_conflicting_row_is_skipped_unless_confirmed(db_session, session_factory, tmp_path):
    exported_at = utcnow()
    product = make_product(db_session, name="Original", code="C-1")
    product.updated_at = exported_at + timedelta(minutes=5)
    db_session.commit()

    content = build_file(
        [{"id": product.id, "code": "C-1", "name": "From file", "type": "Product"}],
        exported_at=exported_at,
    )
    job = analyze(db_session, session_factory, content, tmp_path)
    apply_job(db_session, session_factory, job, apply_conflicts=False)

    db_session.refresh(product)
    assert job.skipped_count == 1
    assert product.name == "Original"


def test_conflicting_row_is_applied_when_confirmed(db_session, session_factory, tmp_path):
    exported_at = utcnow()
    product = make_product(db_session, name="Original", code="C-1")
    product.updated_at = exported_at + timedelta(minutes=5)
    db_session.commit()

    content = build_file(
        [{"id": product.id, "code": "C-1", "name": "From file", "type": "Product"}],
        exported_at=exported_at,
    )
    job = analyze(db_session, session_factory, content, tmp_path)
    apply_job(db_session, session_factory, job, apply_conflicts=True)

    db_session.refresh(product)
    assert job.updated_count == 1
    assert product.name == "From file"


# ----------------------------
# Endpoint behaviour
# ----------------------------
def test_upload_rejects_non_xlsx(client, auth_headers, employee):
    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.csv", b"a,b", "text/csv")},
        headers=auth_headers,
    )

    assert response.status_code == 400
    assert "xlsx" in response.json()["detail"]


def test_upload_rejects_file_over_the_row_limit(client, auth_headers, employee, monkeypatch):
    monkeypatch.setattr(bulk_api, "MAX_DATA_ROWS", 2)
    content = build_file([{"name": f"P{index}", "type": "Product"} for index in range(3)])

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", content, XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 400
    assert "limit is 2" in response.json()["detail"]


def test_an_unanswered_preview_never_blocks_a_new_upload(
    client, auth_headers, db_session, employee, tmp_path
):
    """Abandoning a preview changed nothing, so it must not deadlock the module."""
    abandoned = create_job(db_session, build_file([]), tmp_path, status="awaiting_confirmation")

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", build_file([]), XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 202, response.text

    db_session.expire_all()
    superseded = db_session.get(ProductBulkImports, abandoned.id)
    assert superseded.status == "cancelled"
    assert "Superseded" in superseded.failure_reason


def test_a_running_import_still_blocks_a_new_upload(
    client, auth_headers, db_session, employee, tmp_path
):
    running = create_job(db_session, build_file([]), tmp_path, status="applying")
    running.heartbeat_at = utcnow()
    db_session.commit()

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", build_file([]), XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 409
    detail = response.json()["detail"]
    assert "applying" in detail
    # The message must name the way out, not just the wall.
    assert "automatically" in detail


def test_truncated_upload_says_how_big_the_file_was(client, auth_headers, employee):
    """A download aborted mid-flight lands here; the size makes that obvious."""
    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", b"", XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 400
    assert "empty" in response.json()["detail"]


def test_truncated_xlsx_says_the_download_was_cut_short(client, auth_headers, db_session, employee):
    """A real export cut mid-transfer still starts with PK but has no zip directory."""
    make_product(db_session, name="Panel")
    exported = client.get("/products/export", headers=auth_headers)
    truncated = exported.content[: len(exported.content) // 4]

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", truncated, XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "incomplete" in detail
    assert "Export the catalog again" in detail
    # Zip internals are an implementation detail, not something to hand the user.
    assert "zip" not in detail.lower()


def test_excel_online_error_page_is_named_for_what_it_is(client, auth_headers, employee):
    """Excel Online saves its own failure under the .xlsx name; say so plainly."""
    page = (
        b"<html><head></head><body><script>window.ewaResult = "
        b'{"Errors":[{"MessageIdName":"InternalErrorEwr"}]}</script></body></html>'
    )

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", page, XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "Excel Online" in detail
    assert "desktop app" in detail


@pytest.mark.parametrize(
    ("content", "expected"),
    [
        (b"<!DOCTYPE html><html></html>", "web page"),
        (b'{"detail":"nope"}', "server response"),
        (b"\xd0\xcf\x11\xe0legacy", "legacy .xls"),
    ],
)
def test_upload_names_what_the_file_actually_is(client, auth_headers, employee, content, expected):
    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", content, XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 400
    assert expected in response.json()["detail"]


def test_import_requires_create_and_edit(client, auth_headers, permission_override):
    permission_override("products", {"can_create"})

    response = client.post(
        "/products/bulk-imports",
        files={"file": ("products.xlsx", build_file([]), XLSX_MEDIA_TYPE)},
        headers=auth_headers,
    )

    assert response.status_code == 403


def test_confirm_rejects_a_job_that_is_not_awaiting_confirmation(
    client, auth_headers, db_session, employee, tmp_path
):
    job = create_job(db_session, build_file([]), tmp_path, status="analyzing")

    response = client.post(
        f"/products/bulk-imports/{job.id}/confirm",
        json={"create_missing_taxonomy": False, "apply_conflicts": False},
        headers=auth_headers,
    )

    assert response.status_code == 409


def test_history_lists_who_ran_the_import(client, auth_headers, db_session, employee, tmp_path):
    create_job(db_session, build_file([]), tmp_path, status="completed", created_count=3)

    response = client.get("/products/bulk-imports", headers=auth_headers)

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["total"] == 1
    assert payload["items"][0]["created_count"] == 3
    assert payload["items"][0]["creator"]["email"] == "test@example.com"


def test_history_hides_cancelled_imports(client, auth_headers, db_session, employee, tmp_path):
    """A cancelled import changed nothing, so it is not part of the record."""
    create_job(db_session, build_file([]), tmp_path, status="completed", created_count=2)
    create_job(db_session, build_file([]), tmp_path, status="failed", failure_reason="boom")
    create_job(db_session, build_file([]), tmp_path, status="cancelled")

    payload = client.get("/products/bulk-imports", headers=auth_headers).json()

    assert payload["total"] == 2
    assert {item["status"] for item in payload["items"]} == {"completed", "failed"}


def test_interrupted_job_is_expired_by_the_stale_heartbeat(
    client, auth_headers, db_session, employee, tmp_path
):
    job = create_job(db_session, build_file([]), tmp_path, status="applying")
    job.heartbeat_at = utcnow() - timedelta(hours=1)
    db_session.commit()

    response = client.get(f"/products/bulk-imports/{job.id}", headers=auth_headers)

    assert response.status_code == 200, response.text
    assert response.json()["status"] == "failed"
    assert "restarted" in response.json()["failure_reason"]


def test_unreadable_file_fails_the_job_with_a_reason(db_session, session_factory, tmp_path):
    job = create_job(db_session, b"this is not a workbook", tmp_path)

    run_analysis(job.id, session_factory)

    db_session.refresh(job)
    assert job.status == "failed"
    assert job.failure_reason


def test_full_round_trip_through_the_api(client, auth_headers, db_session, employee, taxonomy):
    """Export, edit the real file, upload it, confirm — end to end over HTTP."""
    from main import app

    import tests.conftest as conftest
    from bd.dependencies import get_session_factory

    class _Factory:
        def __enter__(self):
            self.session = Session(bind=conftest._test_engine.connect())
            return self.session

        def __exit__(self, *_args):
            self.session.close()

    app.dependency_overrides[get_session_factory] = lambda: _Factory
    try:
        family, category = taxonomy
        kept = make_product(db_session, name="Kept", code="RT-1", family_id=family.id)
        edited = make_product(db_session, name="Old Price", code="RT-2", family_id=family.id)

        exported = client.get("/products/export", headers=auth_headers)
        assert exported.status_code == 200, exported.text

        # Edit the exported file the way a user would: change a price, drop a
        # row entirely, and append a brand new product with no id.
        workbook = load_workbook(io.BytesIO(exported.content))
        sheet = workbook["Products"]
        header = [cell.value for cell in sheet[1]]
        id_column = header.index("id") + 1
        name_column = header.index("name") + 1
        price_column = header.index("unit_price") + 1

        for row in range(sheet.max_row, 1, -1):
            if sheet.cell(row=row, column=id_column).value == kept.id:
                sheet.delete_rows(row)
            elif sheet.cell(row=row, column=id_column).value == edited.id:
                sheet.cell(row=row, column=price_column).value = 99.5
        sheet.append([None] * (name_column - 1) + ["Appended Product"])
        appended_row = sheet.max_row
        sheet.cell(row=appended_row, column=header.index("type") + 1).value = "Service"

        buffer = io.BytesIO()
        workbook.save(buffer)

        upload = client.post(
            "/products/bulk-imports",
            files={"file": ("edited.xlsx", buffer.getvalue(), XLSX_MEDIA_TYPE)},
            headers=auth_headers,
        )
        assert upload.status_code == 202, upload.text
        import_id = upload.json()["id"]

        preview = client.get(f"/products/bulk-imports/{import_id}", headers=auth_headers).json()
        assert preview["status"] == "awaiting_confirmation", preview
        assert preview["create_count"] == 1
        assert preview["update_count"] == 1

        confirmed = client.post(
            f"/products/bulk-imports/{import_id}/confirm",
            json={"create_missing_taxonomy": False, "apply_conflicts": False},
            headers=auth_headers,
        )
        assert confirmed.status_code == 202, confirmed.text

        result = client.get(f"/products/bulk-imports/{import_id}", headers=auth_headers).json()
        assert result["status"] == "completed", result
        assert result["created_count"] == 1
        assert result["updated_count"] == 1

        db_session.expire_all()
        assert float(db_session.get(Products, edited.id).unit_price) == pytest.approx(99.5)
        # The row deleted from the file must survive untouched.
        assert db_session.get(Products, kept.id).name == "Kept"
        assert db_session.exec(select(Products).where(Products.name == "Appended Product")).first()
    finally:
        app.dependency_overrides.pop(get_session_factory, None)


def test_import_never_writes_stock_quantity(db_session, session_factory, tmp_path):
    product = make_product(db_session, name="Stocked", code="S-1", stock_quantity=42)

    content = build_file(
        [{"id": product.id, "code": "S-1", "name": "Renamed", "type": "Product"}],
        columns=[*COLUMNS, "stock_quantity"],
    )
    job = analyze(db_session, session_factory, content, tmp_path)
    apply_job(db_session, session_factory, job)

    db_session.refresh(product)
    assert product.name == "Renamed"
    assert product.stock_quantity == 42
